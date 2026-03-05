"""
Build the product catalog JSON by combining:
- Excel price list (codes, descriptions, prices)
- PDF brochure (categories, page structure)
- Image mapping (extracted product images)

Groups size variants under parent products and assigns categories.
"""

import os
import re
import json
import openpyxl

EXCEL_PATH = os.path.join(os.path.dirname(__file__), '..', 'BROCHURE & PRICELIST', 'Persimmon-Product-List-v26 (OSAD).xlsx')
IMAGE_MAPPING_PATH = os.path.join(os.path.dirname(__file__), 'image_mapping.json')
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), '..', 'shop', 'data', 'catalog.json')

# Size suffix pattern: /E, /F, /M, /Y, /W, /H, etc. (sometimes with extra chars)
SIZE_SUFFIX = re.compile(r'^(.+?)/([A-Z]+\d*)$')

# Map size codes to human-readable names
SIZE_NAMES = {
    'E': 'Extra Small',
    'F': 'Small',
    'M': 'Medium',
    'Y': 'Large',
    'W': 'Extra Large',
    'X': 'XXL',
    'H': 'Banner',
}

# Category definitions: maps PDF page ranges to categories
# Built from PDF content analysis
CATEGORY_DEFINITIONS = [
    {
        'name': 'Site Setup Pack',
        'slug': 'site-setup-pack',
        'description': 'Save money and time when ordering signage with these great value temporary site sign packs.',
        'code_prefixes': ['PA115'],
        'keywords': ['site setup pack', 'sign pack'],
    },
    {
        'name': 'Environmental Signs',
        'slug': 'environmental-signs',
        'description': 'Help protect the environment and raise awareness with temporary environmental and waste signs.',
        'code_prefixes': ['PA518', 'PA647', 'PA521', 'PA646', 'PA52', 'PA54', 'PA55', 'PA530', 'PA531', 'PA532', 'PA533', 'PA535', 'PA536', 'PA537', 'PA538'],
        'keywords': ['waste', 'recycl', 'hazardous waste', 'environment', 'inert', 'topsoil', 'timber', 'metals', 'glass', 'general waste', 'road sweeper'],
    },
    {
        'name': 'Site Entrance Signs',
        'slug': 'site-entrance-signs',
        'description': 'Ensure site entrances are clearly marked for workers and visitors to your construction site.',
        'code_prefixes': ['PCFA107', 'PCFA', 'PA681', 'PCFCCS18', 'PCF167', 'PCF151', 'PCF971', 'PCF970', 'PCF384', 'PCF40', 'PCF329'],
        'keywords': ['site entrance', 'visitors must report', 'unauthorised entry'],
    },
    {
        'name': 'Information Signs',
        'slug': 'information-signs',
        'description': 'Display essential construction site information around the site entrance, offices and welfare facilities.',
        'code_prefixes': ['PCF128', 'PCF52', 'PCF414', 'PCF127', 'PCF15', 'PCF25', 'PCFHL', 'PA116', 'PA128PCF', 'PA129PCF', 'PA130PCF'],
        'keywords': ['assembly point', 'no smoking', 'vaporiz', 'information board', 'company information', 'construction information', 'visitors & deliveries', 'works access'],
    },
    {
        'name': 'P.P.E. Signs',
        'slug': 'ppe-signs',
        'description': 'Personal protective equipment signs help enforce the use of mandatory safety equipment.',
        'code_prefixes': ['PA685', 'PCF106', 'PCF961', 'PCF962', 'PCF05', 'PCF129', 'PA120', 'PA121', 'PA122', 'PA123', 'PA49'],
        'keywords': ['ppe', 'safety helmet', 'eye protection', 'gloves must', 'hi-vis', 'high visibility', 'safety footwear', 'ear protection', 'hard hat', 'hard hats'],
    },
    {
        'name': 'Hazard Signs',
        'slug': 'hazard-signs',
        'description': 'Help prevent accidents by warning about potential hazards and dangers on your construction site.',
        'code_prefixes': ['PCF297', 'PCF293', 'PCF294', 'PCF298', 'PCF328', 'PCF323', 'PA46', 'PA47', 'PA48'],
        'keywords': ['danger', 'hazard', 'caution', 'warning', 'deep excavation', 'men working overhead', 'asbestos', 'high voltage', 'underground live', 'thumbs up'],
    },
    {
        'name': 'Fire Signs',
        'slug': 'fire-signs',
        'description': 'Temporary fire exit and fire equipment signs to aid in the event of an emergency.',
        'code_prefixes': ['PCF701', 'PCF708', 'PCF703', 'PCF704', 'PCF709', 'F08B', 'PA50', 'PA51'],
        'keywords': ['fire exit', 'fire point', 'fire extinguisher', 'fire assembly', 'fire action'],
    },
    {
        'name': 'Pedestrian Signs',
        'slug': 'pedestrian-signs',
        'description': 'Guide pedestrians safely around your construction site with temporary signs.',
        'code_prefixes': ['PCF375', 'PCF376', 'PCF346', 'PCF344', 'PCF345', 'PCF326', 'PCF324', 'PCF325', 'PA45'],
        'keywords': ['pedestrian'],
    },
    {
        'name': 'Security Signs',
        'slug': 'security-signs',
        'description': 'Help protect your premises from intruders or warn of security measures such as CCTV.',
        'code_prefixes': ['PCF65', 'PCF03', 'PCF04', 'PCF20', 'PCF511', 'PCF79', 'PA118', 'PA41', 'PA124'],
        'keywords': ['cctv', 'security', 'trespasser', 'keep out', 'guard dog', 'intruder', 'unauthorised persons', 'security on patrol'],
    },
    {
        'name': 'Health Signs',
        'slug': 'health-signs',
        'description': 'Display signage around welfare facilities to ensure essential areas can be easily located.',
        'code_prefixes': ['PCF91', 'PCF463', 'PCF462', 'PCF712', 'PCF464', 'PA125'],
        'keywords': ['first aid', 'toilet', 'canteen', 'drying room', 'wash', 'welfare', 'drinking water', 'nearest first aid'],
    },
    {
        'name': 'Site Marking',
        'slug': 'site-marking',
        'description': 'Identify and mark out cables, pipes and underground services with highly visible temporary signs.',
        'code_prefixes': ['MTP', 'PCFMK', 'PCFSTPK'],
        'keywords': ['marker tape', 'service marker', 'stencil', 'room i.d'],
    },
    {
        'name': 'Simple Safety Signs',
        'slug': 'simple-safety-signs',
        'description': 'Help children and visitors easily understand health and safety signs on your construction site.',
        'code_prefixes': ['PCFB', 'PCFS0', 'PCFSN'],
        'keywords': [],
    },
    {
        'name': 'Working at Height',
        'slug': 'working-at-height',
        'description': 'Scaffolding and ladder safety products to keep your workforce safe whilst working at height.',
        'code_prefixes': ['LTK', 'SIK', 'TTK', 'PCFSB', 'PCFSF', 'PA686', 'PA687', 'PA119'],
        'keywords': ['scaffold', 'ladder', 'working at height', 'scaf-fold'],
    },
    {
        'name': 'Traffic Signs',
        'slug': 'traffic-signs',
        'description': 'Ensure correct temporary site traffic signs are in place to minimise risk of traffic accidents.',
        'code_prefixes': ['PCF251', 'PCF24', 'PCF252', 'PCF69', 'PCF35', 'PCF33', 'PCFTM', 'PA117', 'PA126'],
        'keywords': ['site traffic', 'speed limit', 'slow', 'traffic management'],
    },
    {
        'name': 'Parking Signs',
        'slug': 'parking-signs',
        'description': 'Guide motorists safely to designated parking areas on your construction site.',
        'code_prefixes': ['PCF353', 'PCF315', 'PCF350', 'PCF351', 'PCF352', 'PCF311', 'PA40', 'PA42', 'PA43'],
        'keywords': ['parking', 'no parking', 'residents parking'],
    },
    {
        'name': 'Roadworks Signs',
        'slug': 'roadworks-signs',
        'description': 'Guide motorists and pedestrians through temporary roadworks and diversions.',
        'code_prefixes': ['PCFR'],
        'keywords': ['roadwork', 'road ahead', 'diversion'],
    },
    {
        'name': 'Freestanding Signs',
        'slug': 'freestanding-signs',
        'description': 'Freestanding sign frames - easy to pick up and transport, or put away and store when not in use.',
        'code_prefixes': ['AS0', 'PCFSG', 'PCFDI', 'PCFPU'],
        'keywords': ['a-stand', 'freestanding', 'squarecade'],
    },
    {
        'name': 'Display Boards',
        'slug': 'display-boards',
        'description': 'Large display boards for displaying essential construction site information.',
        'code_prefixes': ['PCFDW', 'PCFCCS20', 'PCFCCS51', 'PCFCF', 'PCFPB', 'DWA', 'PCFPL'],
        'keywords': ['dry wipe', 'display board', 'pinboard', 'clipframe', 'magnetic board'],
    },
    {
        'name': 'Notice Boards',
        'slug': 'notice-boards',
        'description': 'Noticeboards for organising essential construction site forms and information.',
        'code_prefixes': ['PA97', 'PA96', 'PA99', 'PA100', 'PA101', 'PA102', 'PA104', 'PA105', 'PA106', 'PA107', 'PA108'],
        'keywords': ['notice board', 'information board', 'score board', 'check list board', 'crane information', 'daily hazard', 'call off'],
    },
    {
        'name': 'Considerate Site Signs',
        'slug': 'considerate-site-signs',
        'description': 'Signs designed to encourage best practice, focusing on respecting the community and environment.',
        'code_prefixes': ['PA109', 'PA110', 'PA111', 'PA112', 'PA113', 'PA114', 'PA602', 'PA89', 'PA90', 'PA93', 'PA94', 'PA288'],
        'keywords': ['we aim to please', 'considerate', 'apologise for any inconvenience', 'for sale', 'main compound board', 'safety concern'],
    },
    {
        'name': 'Posters',
        'slug': 'posters',
        'description': 'Health and safety posters providing helpful information and guidance for employees.',
        'code_prefixes': ['PCFS103', 'PCFS104', 'PCFS105', 'PCFS106', 'PCFS88', 'PCFS38', 'PCFS86', 'PCFS89', 'PCFS85', 'PCFS84', 'PCFS83', 'PCFS87', 'PCFS91', 'PCFCCS', 'PCFSUN'],
        'keywords': ['poster', 'sun safety'],
    },
    {
        'name': 'Prestige Signs',
        'slug': 'prestige-signs',
        'description': 'Give your work environment a professional look with high quality prestige signs.',
        'code_prefixes': ['PCFPR'],
        'keywords': ['engraved', 'anodised aluminium sign', 'door disc'],
    },
    {
        'name': 'Hoarding Signs',
        'slug': 'hoarding-signs',
        'description': 'Advertise your brand with high quality marketing signs and site perimeter banners.',
        'code_prefixes': ['PCFHS', 'PCFHSF', 'PCFCCS06', 'PCFHSE', 'PCFHSEM'],
        'keywords': ['hoarding', 'mesh banner'],
    },
    {
        'name': 'Finished Home Signs',
        'slug': 'finished-home-signs',
        'description': 'Add the finishing touch to your development with great-value finished home signs.',
        'code_prefixes': ['PCFFH', 'SNC', 'PCFBA'],
        'keywords': ['welcome door', 'floor protection', 'toilet seal', 'manifestation', 'welcome floor mat', 'utility sticker', 'parking disc', 'we called today', 'street name'],
    },
    {
        'name': 'Custom Signs',
        'slug': 'custom-signs',
        'description': 'Create your perfect construction sign with pre-designed custom sign templates.',
        'code_prefixes': ['PCF995', 'PCF996', 'PCF997', 'PCF998', 'PCF999', 'PCFMK95', 'PCFMK97', 'PCFMK98'],
        'keywords': ['custom'],
    },
    {
        'name': 'Extras & Accessories',
        'slug': 'extras-accessories',
        'description': 'Sign fixings, sign posts and extra accessories for displaying your signs.',
        'code_prefixes': ['XR', 'TMA', 'PCFLFB', 'PCFPFB', 'PCFSAF'],
        'keywords': ['post cap', 'mirror', 'sign ties', 'clip', 'bracket', 'lanyard', 'holder'],
    },
]


def parse_excel():
    """Parse all products from the Excel price list."""
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['Product List']

    products = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        code = row[0]
        description = row[1]
        retail_price = row[2]
        persimmon_price = row[3]

        if not code or not description:
            continue

        products.append({
            'code': str(code).strip(),
            'description': str(description).strip(),
            'retailPrice': float(retail_price) if retail_price else 0,
            'price': float(persimmon_price) if persimmon_price else 0,
        })

    return products


def get_base_code(code):
    """Extract base code from a product code."""
    match = SIZE_SUFFIX.match(code)
    if match:
        return match.group(1), match.group(2)
    return code, None


def extract_size_from_description(description):
    """Extract dimensions from description like '300x400mm'."""
    match = re.search(r'(\d+x\d+mm)', description)
    if match:
        return match.group(1)
    return None


def extract_material_from_description(description):
    """Extract material type from description."""
    materials = [
        '10mm Correx', '4mm Correx', '5mm Foamex', '2mm Correx',
        'PVC Banner', 'Magnetic Dry Wipe Board', 'Semi-rigid PVC',
        'Zintec', 'Anodised Aluminium', 'Stainless Steel',
        'Self-Adhesive Vinyl', 'Dibond', 'Acrylic',
        'SCAF-FOLD', 'PVC Mesh Banner',
    ]
    desc_lower = description.lower()
    for mat in materials:
        if mat.lower() in desc_lower:
            return mat
    # Check parenthetical material
    paren_match = re.search(r'\(([^)]+)\)\s*$', description)
    if paren_match:
        return paren_match.group(1)
    return None


def extract_product_name(description):
    """Extract the human-readable product name from description (removing size and material)."""
    # Remove leading dimensions
    name = re.sub(r'^\d+x\d+mm\s*', '', description)
    # Remove trailing parenthetical material info
    name = re.sub(r'\s*\([^)]*\)\s*$', '', name)
    # Remove size descriptors
    name = re.sub(r'\s*\(A\d\)\s*', ' ', name)
    # Clean up
    name = name.strip()
    # Remove "Site address to read:" and similar trailing instructions
    name = re.sub(r'\s*Site address to read:.*$', '', name)
    return name if name else description


def categorize_product(code, description=''):
    """Assign a product to a category based on its code prefix, with keyword fallback."""
    # First try code prefix match (most reliable)
    for cat_def in CATEGORY_DEFINITIONS:
        for prefix in cat_def['code_prefixes']:
            if code.startswith(prefix):
                return cat_def['slug']

    # Fallback: keyword match on description
    desc_lower = description.lower()
    for cat_def in CATEGORY_DEFINITIONS:
        for keyword in cat_def.get('keywords', []):
            if keyword.lower() in desc_lower:
                return cat_def['slug']

    return None


def build_catalog():
    """Build the full product catalog."""
    # Load image mapping
    image_mapping = {}
    if os.path.exists(IMAGE_MAPPING_PATH):
        with open(IMAGE_MAPPING_PATH, 'r', encoding='utf-8') as f:
            image_mapping = json.load(f)

    # Parse Excel products
    products = parse_excel()
    print(f"Loaded {len(products)} products from Excel")

    # Group variants under base codes
    product_groups = {}
    for product in products:
        base_code, size_suffix = get_base_code(product['code'])

        if base_code not in product_groups:
            product_groups[base_code] = {
                'baseCode': base_code,
                'variants': [],
                'name': extract_product_name(product['description']),
            }

        product_groups[base_code]['variants'].append({
            'code': product['code'],
            'size': extract_size_from_description(product['description']),
            'material': extract_material_from_description(product['description']),
            'description': product['description'],
            'price': round(product['retailPrice'], 2),
            'persimmonPrice': round(product['price'], 2),
            'sizeSuffix': size_suffix,
        })

    print(f"Grouped into {len(product_groups)} base products")

    # Assign images
    with_image = 0
    for base_code, group in product_groups.items():
        if base_code in image_mapping:
            group['image'] = image_mapping[base_code]
            with_image += 1
        else:
            group['image'] = None
    print(f"Products with images: {with_image}/{len(product_groups)}")

    # Build categories
    categories = []
    categorized_count = 0
    uncategorized = []

    for cat_def in CATEGORY_DEFINITIONS:
        cat = {
            'name': cat_def['name'],
            'slug': cat_def['slug'],
            'description': cat_def['description'],
            'products': [],
        }
        categories.append(cat)

    # Assign products to categories
    for base_code, group in product_groups.items():
        # Use first variant description for keyword matching
        first_desc = group['variants'][0]['description'] if group['variants'] else ''
        slug = categorize_product(base_code, first_desc)
        if slug:
            for cat in categories:
                if cat['slug'] == slug:
                    cat['products'].append(group)
                    categorized_count += 1
                    break
        else:
            uncategorized.append(base_code)

    # Sort products within categories by base code
    for cat in categories:
        cat['products'].sort(key=lambda p: p['baseCode'])
        cat['productCount'] = len(cat['products'])

    # Put uncategorized products in "Other Signs" category
    if uncategorized:
        other_cat = {
            'name': 'Other Signs',
            'slug': 'other-signs',
            'description': 'Additional signage products available for your construction site.',
            'products': [],
        }
        for base_code in uncategorized:
            if base_code in product_groups:
                other_cat['products'].append(product_groups[base_code])
        other_cat['products'].sort(key=lambda p: p['baseCode'])
        other_cat['productCount'] = len(other_cat['products'])
        categories.append(other_cat)

    # Remove empty categories
    categories = [c for c in categories if c['products']]

    print(f"Categorized: {categorized_count}/{len(product_groups)}")
    print(f"Uncategorized: {len(uncategorized)}")
    if uncategorized[:20]:
        print(f"  Sample uncategorized codes: {uncategorized[:20]}")

    catalog = {
        'categories': categories,
        'totalProducts': len(product_groups),
        'totalVariants': len(products),
    }

    # Ensure output directory exists
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)

    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(catalog, f, indent=2, ensure_ascii=False)

    print(f"\nCatalog saved to: {OUTPUT_PATH}")
    print(f"Categories: {len(categories)}")
    for cat in categories:
        print(f"  {cat['name']}: {cat['productCount']} products")

    return catalog


if __name__ == '__main__':
    build_catalog()
